from googleapiclient.discovery import build
from googleapiclient.http import MediaInMemoryUpload
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl.styles import Alignment
import pickle
import os
import io


def upload_to_google_drive(df, folder_id, output_filename):
    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    TOKEN_FILE = r'C:\Users\wuhan\.gcp\token.pickle'
    CREDENTIALS_FILE = r'C:\Users\wuhan\.gcp\drive_auth.json'

    creds = None
    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, 'rb') as f:
            creds = pickle.load(f)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, 'wb') as f:
            pickle.dump(creds, f)

    service = build('drive', 'v3', credentials=creds)

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')

    from openpyxl import load_workbook
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    media = MediaInMemoryUpload(buffer.read(), mimetype=mimetype)

    # check if file already exists in the folder
    query = f"name='{output_filename}' and '{folder_id}' in parents and trashed=false"
    existing = service.files().list(q=query, fields='files(id)').execute().get('files', [])

    if existing:
        file_id = existing[0]['id']
        uploaded = service.files().update(fileId=file_id, media_body=media, fields='id').execute()
        print(f"Overwritten file ID: {uploaded.get('id')}")
    else:
        file_metadata = {'name': output_filename, 'parents': [folder_id]}
        uploaded = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"Uploaded file ID: {uploaded.get('id')}")


# ============================================================================
# TEST NEW AUTO EVALUATION
# ============================================================================
from pydantic import BaseModel, Field
import pandas as pd
import yaml
import os
import json
import asyncio
import psycopg2
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from groq import RateLimitError

from langchain_groq import ChatGroq
from langchain.prompts import PromptTemplate
from langchain.output_parsers import PydanticOutputParser
from langchain.output_parsers import OutputFixingParser

_here = os.path.dirname(os.path.abspath(__file__))
with open(os.path.join(_here, 'eval_prompts.yaml'), 'r') as file:
    prompt_versions = yaml.safe_load(file)

eval_llm = ChatGroq(
    groq_api_key=os.environ["GROQ_TOKEN"],
    model_name="openai/gpt-oss-20b", 
    temperature=0.7
)

@retry(
    retry=retry_if_exception_type(RateLimitError),
    wait=wait_exponential(multiplier=1, min=2, max=30),
    stop=stop_after_attempt(5)
)
async def _invoke_with_retry(chain, inputs):
    return await chain.ainvoke(inputs)


def get_eval_input(db_url, config_hash):
    conn = psycopg2.connect(
        host=db_url.host,
        port=db_url.port,
        dbname=db_url.database,
        user=db_url.username,
        password=db_url.password,
    )
    cur = conn.cursor()
    cur.execute("SELECT output FROM existing_rag_output WHERE config_hash = %s", 
                (config_hash,))
    row = cur.fetchone()
    json_results = row[0]
    cur.close()
    conn.close()

    records = []
    for item in json_results:
        record = {
            'query': item['question'],
            'ai_answer': item['ai_answer'],
            'referenced_answer': item['expected_answer'],
            'retrieved_content': ''.join(content_dct['content'] for content_dct in item['retrieved_lst']),
        }
        records.append(record)
    return pd.DataFrame(records)


async def eval_one_config(config_hash, db_url, rag_df):    
    input_df = get_eval_input(db_url, config_hash)
    input_df = pd.merge(input_df, rag_df[['question', 'context']], 
                        left_on='query', right_on='question')
    input_df.drop(columns=['question'], inplace=True)

    retrieval_quality = await get_retrieval_quality_output_async(input_df, eval_llm,
                                                            prompt_versions['rq_prompt_template'])
    retrieval_quality['same_context'] = retrieval_quality['retrieved_content'] == retrieval_quality['context']
    
    answer_quality = await get_answer_quality_output_async(input_df, eval_llm,
                                                            prompt_versions['aq_prompt_template'])

    return config_hash, retrieval_quality, answer_quality


async def run_auto_eval(config_hashes, db_url, rag_df):
    results = await asyncio.gather(*[
        eval_one_config(config_hash, db_url, rag_df)
        for config_hash in config_hashes
    ])

    return {
         config_hash: {"retrieval_quality_df": retrieval_quality,
                        "answer_quality_df": answer_quality}
         for config_hash, retrieval_quality, answer_quality in results
     }


# ------------------------------------------ RETRIEVAL QUALITY ------------------------------------------ #
class RetrievalQuality(BaseModel):
    score: int = Field(description="""Score with:
                - Only generate the score as -1, 0 or 1 or 2 or 3 or 4
                - Scoring as -1: if the RETRIEVED CONTENT is much more relevant to the USER QUERY than the CONTEXT
                - Scoring as 0: if the RETRIEVED CONTENT is completely irrelevant to the USER QUERY
                - If the CONTEXT is strongly relevant to the USER QUERY:
                    - Scoring as 1: if the RETRIEVED CONTENT is relevant to the USER QUERY but doesn't contain any critical information from the CONTEXT
                    - Scoring as 2: if the RETRIEVED CONTENT is relevant to the USER QUERY but only partially contains critical information from the CONTEXT
                    - Scoring as 3: if the RETRIEVED CONTENT is relevant to USER QUERY and contains all the critical information from the CONTEXT
                    - Scoring as 4: if the RETRIEVED CONTENT is relevant to USER QUERY and contains more critical information than the CONTEXT that can help answer the USER QUERY
            """)
    reasoning: str = Field(description="Reasoning for the given score.")


async def evaluate_retrieval_quality_async(llm, user_query, context, retrieved_content, rq_prompt_template):
    base_parser = PydanticOutputParser(pydantic_object=RetrievalQuality)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=rq_prompt_template,
        input_variables=["user_query", "context", "retrieved_content"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query,
        "context": context,
        "retrieved_content": retrieved_content
    })
    return result


async def process_retrieval_quality_record_async(llm, record, rq_prompt_template):
    eval_result = await evaluate_retrieval_quality_async(
        llm,
        record['query'],
        record['context'],
        record['retrieved_content'],
        rq_prompt_template
    )
    record['retrieval_quality_score'] = eval_result.score
    record['rq_reasoning'] = eval_result.reasoning
    return record


async def get_retrieval_quality_output_async(input_df, llm, rq_prompt_template, concurrency=2):
    sem = asyncio.Semaphore(concurrency)  # Semaphore to throttle concurrency

    async def sem_task(record):
        async with sem:
            return await process_retrieval_quality_record_async(llm, record, rq_prompt_template)

    input_records = input_df.to_dict(orient='records')
    tasks = [sem_task(record) for record in input_records]
    output_lst = await asyncio.gather(*tasks)
    output_df = pd.DataFrame(output_lst)
    return output_df
# ------------------------------------ QUERY QUALITY ------------------------------------ #


# ------------------------------------ ANSWER QUALITY ------------------------------------ #
class AnswerQuality(BaseModel):
    score: int = Field(description="""Score with:
    - Only generate the score as -1, 0 or 1 or 2 or 3 or 4
    - Scoring as -1: if the AI's ANSWER is much more relevant to the USER QUERY than the REFERENCED ANSWER
    - Scoring as 0: if the AI's ANSWER is completely irrelevant to the USER QUERY
    - If the REFERENCED ANSWER is strongly relevant to the USER QUERY:
        - Scoring as 0: if AI's ANSWER is significantly conflict with the REFERENCED ANSWER
        - Scoring as 1: if the AI's ANSWER is relevant to the USER QUERY but doesn't contain any critical information from the REFERENCED ANSWER
        - Scoring as 2: if the AI's ANSWER is relevant to the USER QUERY but only partially contains critical information from the REFERENCED ANSWER
        - Scoring as 3: if the AI's ANSWER is relevant to USER QUERY and contains all the critical information from the REFERENCED ANSWER
        - Scoring as 4: if the AI's ANSWER is relevant to USER QUERY and contains more critical information than the REFERENCED ANSWER that can help answer the USER QUERY
    """)
    reasoning: str = Field(description="Reasoning for the given score.")


async def evaluate_answer_quality_async(llm, user_query, ai_answer, referenced_answer, aq_prompt_template):
    base_parser = PydanticOutputParser(pydantic_object=AnswerQuality)
    output_parser = OutputFixingParser.from_llm(parser=base_parser, llm=llm)
    prompt = PromptTemplate(
        template=aq_prompt_template,
        input_variables=["user_query", "ai_answer", "referenced_answer"],
        partial_variables={"format_instructions": output_parser.get_format_instructions()},
    )
    chain = prompt | llm | output_parser
    result = await _invoke_with_retry(chain, {
        "user_query": user_query,
        "ai_answer": ai_answer,
        "referenced_answer": referenced_answer
    })
    return result


async def process_answer_quality_record_async(llm, record, aq_prompt_template):
    eval_result = await evaluate_answer_quality_async(
        llm,
        record['query'],
        record['ai_answer'],
        record['referenced_answer'],
        aq_prompt_template
    )
    record['answer_quality_score'] = eval_result.score
    record['aq_reasoning'] = eval_result.reasoning
    return record


async def get_answer_quality_output_async(input_df, llm, aq_prompt_template, concurrency=2):
    sem = asyncio.Semaphore(concurrency)  # Semaphore to throttle concurrency

    async def sem_task(record):
        async with sem:
            return await process_answer_quality_record_async(llm, record, aq_prompt_template)

    input_records = input_df.to_dict(orient='records')
    tasks = [sem_task(record) for record in input_records]
    output_lst = await asyncio.gather(*tasks)
    output_df = pd.DataFrame(output_lst)
    return output_df
# ------------------------------------ ANSWER QUALITY ------------------------------------ #